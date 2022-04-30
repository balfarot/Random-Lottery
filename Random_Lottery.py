#List of Name out of a Excel Sheet with Graphical User Interface

from tkinter import *
from tkinter.ttk import *

from openpyxl import workbook, load_workbook
import random

root = Tk()
root.title('Random Lottery')
root.config(bg = 'pink')
root. geometry('200x200')
root.resizable(0,0) #people can't expand

def randomname(event):
    wb = load_workbook('hello.xlsx')
    ws = wb.active
    rangeline = ws['A2':'A19']
    name = []
    for items in rangeline:
            for subitems in items:
                name.append(subitems.value)
    computer_action = random.choice(name)
    print('The computer randomly chose ' + computer_action)

    updateDisplay(computer_action)

def updateDisplay(abc):
    displayvariable.set(abc)

button_1 = Button(root, text = "Random Lottery Name")
button_1.bind('<Button-1>', randomname)
button_1.pack()
displayvariable = StringVar()
displaylabel = Label(root, textvariable = displayvariable)
displaylabel.pack()

mainloop()